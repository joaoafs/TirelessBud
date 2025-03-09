import os
import subprocess
import sys
import time
import threading
import tempfile
from tkinter import Tk, Label, Button, Entry, filedialog, messagebox, Toplevel, StringVar, Frame, ttk
from tkinter.font import Font
import shutil
import urllib.request
import json
import webbrowser

# Versão atual do programa
VERSION = "1.0.0"
REPO_URL = "https://github.com/joaoafs/TirelessBud"
RELEASES_URL = "https://github.com/joaoafs/TirelessBud/releases"

# Arquivo de bloqueio para evitar múltiplas instâncias
LOCK_FILE = os.path.join(tempfile.gettempdir(), "tirelessbud_instance.lock")

# Função para obter o caminho do ícone, funcionando tanto em modo de desenvolvimento quanto em executável
def get_icon_path():
    if getattr(sys, 'frozen', False):
        # Se estiver executando como executável
        base_dir = os.path.dirname(sys.executable)
    else:
        # Se estiver executando como script Python
        # O ícone está na pasta principal do projeto, não na pasta pai
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    icon_path = os.path.join(base_dir, "LogoTBud.ico")
    
    # Verifica se o ícone existe
    if os.path.exists(icon_path):
        return icon_path
    else:
        # Tenta encontrar em um caminho alternativo (pasta principal)
        alt_path = os.path.join(os.path.dirname(base_dir), "LogoTBud.ico")
        if os.path.exists(alt_path):
            return alt_path
    return None

# Função para aplicar ícone à janela
def apply_icon(window):
    icon_path = get_icon_path()
    if (icon_path):
        try:
            window.iconbitmap(icon_path)
        except Exception as e:
            print(f"Erro ao aplicar ícone: {e}")
            pass

# Verifica se já existe uma instância do programa rodando
def check_single_instance():
    if os.path.exists(LOCK_FILE):
        try:
            # Verifica se o arquivo de bloqueio é recente (menos de 10 segundos)
            if time.time() - os.path.getmtime(LOCK_FILE) < 10:
                messagebox.showinfo("Programa já em execução", 
                                  "Uma instância do TirelessBud já está em execução.")
                sys.exit(0)
            else:
                # Arquivo de bloqueio muito antigo, pode ter sido de um fechamento incorreto
                os.remove(LOCK_FILE)
        except:
            # Em caso de erro, tenta remover o arquivo de bloqueio
            try:
                os.remove(LOCK_FILE)
            except:
                pass
    
    # Cria o arquivo de bloqueio
    try:
        with open(LOCK_FILE, "w") as f:
            f.write(str(os.getpid()))
    except:
        # Se não conseguir criar o arquivo de bloqueio, continua mesmo assim
        pass
        
    # Registra função para remover o arquivo de bloqueio ao sair
    import atexit
    atexit.register(lambda: os.remove(LOCK_FILE) if os.path.exists(LOCK_FILE) else None)

# Verificação simplificada de atualização para a versão executável
def check_exe_updates_safe(root):
    try:
        # Uma verificação muito simples, apenas abre a janela de atualização manual
        # sem fazer downloads ou reiniciar automaticamente
        req = urllib.request.Request(
            "https://api.github.com/repos/joaoafs/TirelessBud/releases/latest", 
            headers={'User-Agent': 'Python'}
        )
        with urllib.request.urlopen(req, timeout=3) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            if 'tag_name' in data:
                latest_version = data['tag_name'].lstrip('v')
                
                if latest_version > VERSION:
                    show_update_message(root, VERSION, latest_version)
    except:
        # Ignora erros na verificação de atualizações
        pass

# Mostra mensagem simples sobre atualização disponível
def show_update_message(root, current_version, new_version):
    try:
        update_window = Toplevel(root)
        apply_icon(update_window)  # Aplica o ícone à janela de atualização
        update_window.title("Atualização Disponível")
        center_window(update_window, 400, 200)
        
        Label(update_window, text="Há uma nova versão disponível!", 
              font=("Arial", 14, "bold")).pack(pady=15)
        
        Label(update_window, text=f"Versão atual: {current_version}", 
              font=("Arial", 11)).pack(pady=5)
        Label(update_window, text=f"Nova versão: {new_version}", 
              font=("Arial", 11)).pack(pady=5)
        
        Button(update_window, text="Visitar página de releases", 
              command=lambda: [open_browser_to_url(RELEASES_URL), update_window.destroy()], 
              font=("Arial", 11), bg="#007bff", fg="white", width=20).pack(pady=15)
        
        Button(update_window, text="Mais tarde", 
              command=update_window.destroy, font=("Arial", 11), 
              bg="#6c757d", fg="white", width=20).pack(pady=5)
    except:
        # Ignora erros na criação da janela
        pass

# Função para abrir URL no navegador
def open_browser_to_url(url):
    try:
        webbrowser.open_new(url)
    except:
        try:
            os.startfile(url)
        except:
            messagebox.showinfo("Não foi possível abrir o navegador", 
                              f"Por favor, visite manualmente: {url}")

# Função para verificar atualizações disponíveis (versão segura)
def check_for_updates(root):
    check_exe_updates_safe(root)

# Verificação de dependências instaladas
def check_if_packages_installed():
    packages = ["pymupdf", "PyPDF2", "openpyxl", "pillow"]
    missing_packages = []
    
    for package in packages:
        try:
            # Verifica se o pacote está instalado
            subprocess.check_call([sys.executable, "-m", "pip", "show", package], 
                                 stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except subprocess.CalledProcessError:
            # Pacote não está instalado
            missing_packages.append(package)
    
    return missing_packages

# Verificação e instalação de dependências (versão segura para executável)
def check_and_install_dependencies(root):
    # Primeiro verifica se há atualizações disponíveis
    check_for_updates(root)
    
    # Se estamos rodando como executável, não precisamos verificar dependências
    # pois elas já devem estar incluídas no arquivo .exe
    if getattr(sys, 'frozen', False):
        return
    
    # Apenas se estiver rodando como script Python, verificar dependências
    missing_packages = check_if_packages_installed()
    
    if not missing_packages:
        # Todos os pacotes já estão instalados
        return
    
    root.withdraw()  # Esconde a janela principal
    
    loading_window = Toplevel(root)
    loading_window.title("Instalando Dependências")
    center_window(loading_window, 400, 150)
    
    Label(loading_window, text="A instalar dependências necessárias...", font=("Arial", 12)).pack(pady=15)
    
    progress = ttk.Progressbar(loading_window, orient="horizontal", length=300, mode="indeterminate")
    progress.pack(pady=10)
    progress.start(10)
    
    status_var = StringVar()
    status_var.set("Verificando dependências...")
    status_label = Label(loading_window, textvariable=status_var)
    status_label.pack(pady=10)
    
    loading_window.update()
    
    # Função que processa um pacote por vez e agenda a próxima chamada
    def process_next_package(package_list, index=0):
        if index >= len(package_list):
            # Terminou todos os pacotes
            progress.stop()
            status_var.set("Instalação concluída!")
            loading_window.update()
            loading_window.after(1000, loading_window.destroy)
            root.deiconify()  # Mostra a janela principal
            return
        
        package = package_list[index]
        status_var.set(f"A instalar {package}...")
        loading_window.update()
        
        try:
            # Instala o pacote
            subprocess.check_call([sys.executable, "-m", "pip", "install", package],
                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except Exception as e:
            status_var.set(f"Erro ao instalar {package}: {e}")
            loading_window.update()
        
        # Agenda o próximo pacote para ser processado
        loading_window.after(500, lambda: process_next_package(package_list, index + 1))
    
    # Inicia o processamento na thread principal
    loading_window.after(100, lambda: process_next_package(missing_packages))
    
    # Aguarda a conclusão da instalação
    root.wait_window(loading_window)

# Função para centralizar qualquer janela na tela
def center_window(window, width=None, height=None):
    if width and height:
        # Define a geometria com tamanho explícito
        window.geometry(f"{width}x{height}")
        # Força a atualização para aplicar o tamanho
        window.update_idletasks()
    
    # Obtém dimensões da tela
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    
    # Obtém dimensões atuais da janela
    window_width = window.winfo_width()
    window_height = window.winfo_height()
    
    # Se a janela estiver vazia ou muito pequena, força o tamanho mínimo
    if window_width < 200 or window_height < 200:
        window_width = width if width else 700
        window_height = height if height else 500
    
    # Calcula posição x,y para a janela
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    
    # Define a geometria final com posição centralizada
    window.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    # Forçar a atualização para aplicar o tamanho e posição
    window.update()

# Função para mostrar mensagem inicial com efeito fade (versão completa do main.py)
def show_splash_screen(root):
    try:
        splash = Toplevel(root)
        splash.overrideredirect(True)
        center_window(splash, 500, 250)
        splash.configure(bg="#2E3B4E")
        splash.attributes("-alpha", 0.0)
        
        user_name = os.getlogin()
        messages = [("feito por github.com/joaoafs", 3), (f"bem vindo, {user_name}", 3)]
        
        for message, duration in messages:
            label = Label(splash, text=message, font=("Arial", 20, "bold"), fg="white", bg="#2E3B4E")
            label.pack(expand=True)
            
            for i in range(0, 101, 5):
                splash.attributes("-alpha", i / 100)
                splash.update()
                time.sleep(0.03)
            
            time.sleep(duration)
            
            for i in range(100, -1, -5):
                splash.attributes("-alpha", i / 100)
                splash.update()
                time.sleep(0.03)
            
            label.pack_forget()
        
        splash.destroy()
    except Exception as e:
        # Se ocorrer qualquer erro no splash screen, simplesmente continua sem ele
        print(f"Erro ao exibir splash screen: {e}")
        pass

# Função para mostrar progresso de execução
def show_progress_window(action, parent=None):
    progress_window = Toplevel(parent)
    progress_window.title("Progresso")
    center_window(progress_window, 400, 150)
    
    Label(progress_window, text=f"{action}...", font=("Arial", 12, "bold")).pack(pady=15)
    
    progress = ttk.Progressbar(progress_window, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=10)
    
    progress_var = StringVar()
    progress_var.set("A preparar: 0%")
    progress_label = Label(progress_window, textvariable=progress_var, font=("Arial", 10))
    progress_label.pack(pady=10)
    
    progress_window.update()
    return progress_window, progress, progress_var

# Função para abrir a pasta onde o ficheiro foi salvo
def abrir_pasta(filepath):
    folder = os.path.dirname(filepath)
    if os.path.exists(folder):
        os.startfile(folder)

# Função para unir PDFs
def juntar_pdfs(parent=None):
    try:
        files = filedialog.askopenfilenames(
            parent=parent,
            title="Selecione os ficheiros PDF para juntar",
            filetypes=[("Ficheiros PDF", "*.pdf")]
        )
        if not files:
            return
            
        save_path = filedialog.asksaveasfilename(
            parent=parent,
            title="Guardar PDF unido",
            defaultextension=".pdf",
            filetypes=[("Ficheiros PDF", "*.pdf")]
        )
        if not save_path:
            return
            
        # Importa PyPDF2 aqui para garantir que está instalado
        from PyPDF2 import PdfMerger
        
        progress_window, progress_bar, progress_var = show_progress_window("A unir PDFs", parent)
        
        def executar_uniao():
            try:
                merger = PdfMerger()
                for i, file in enumerate(files):
                    merger.append(file)
                    valor = int((i + 1) / len(files) * 100)
                    progress_bar["value"] = valor
                    progress_var.set(f"A unir: {valor}%")
                    progress_window.update()
                
                merger.write(save_path)
                merger.close()
                progress_bar["value"] = 100
                progress_var.set("Concluído: 100%")
                progress_window.update()
                time.sleep(1)
                progress_window.destroy()
                messagebox.showinfo("Sucesso", f"PDF unido guardado em:\n{save_path}")
                abrir_pasta(save_path)
            except Exception as e:
                progress_window.destroy()
                messagebox.showerror("Erro", f"Erro ao unir PDFs: {e}")
        
        threading.Thread(target=executar_uniao, daemon=True).start()
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar a união de PDFs: {e}")

# Função para separar PDF em páginas
def separar_pdf(parent=None):
    try:
        file_path = filedialog.askopenfilename(
            parent=parent,
            title="Selecione o ficheiro PDF para separar",
            filetypes=[("Ficheiros PDF", "*.pdf")]
        )
        if not file_path:
            return
            
        save_dir = filedialog.askdirectory(
            parent=parent,
            title="Selecione a pasta para guardar os PDFs separados"
        )
        if not save_dir:
            return
            
        # Importa PyPDF2 e fitz aqui para garantir que estão instalados
        from PyPDF2 import PdfReader, PdfWriter
        import fitz
        
        progress_window, progress_bar, progress_var = show_progress_window("A separar PDF", parent)
        
        def executar_separacao():
            try:
                base_filename = os.path.splitext(os.path.basename(file_path))[0]
                
                # Usando PyMuPDF (fitz) que é mais rápido
                doc = fitz.open(file_path)
                total_pages = len(doc)
                
                for page_num in range(total_pages):
                    new_doc = fitz.open()
                    new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
                    
                    # Gera um nome único para cada página
                    output_path = os.path.join(save_dir, f"{base_filename}_ex{page_num + 1}.pdf")
                    
                    # Verifica se o arquivo já existe e incrementa o número se necessário
                    counter = 1
                    while os.path.exists(output_path):
                        output_path = os.path.join(save_dir, f"{base_filename}_ex{page_num + 1}_{counter}.pdf")
                        counter += 1
                    
                    new_doc.save(output_path)
                    new_doc.close()
                    
                    # Atualiza progresso
                    valor = int((page_num + 1) / total_pages * 100)
                    progress_bar["value"] = valor
                    progress_var.set(f"A separar: {valor}%")
                    progress_window.update()
                
                doc.close()
                progress_bar["value"] = 100
                progress_var.set("Concluído: 100%")
                progress_window.update()
                time.sleep(1)
                progress_window.destroy()
                messagebox.showinfo("Sucesso", f"{total_pages} páginas extraídas para:\n{save_dir}")
                abrir_pasta(save_dir)
            except Exception as e:
                progress_window.destroy()
                messagebox.showerror("Erro", f"Erro ao separar PDF: {e}")
        
        threading.Thread(target=executar_separacao, daemon=True).start()
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar a separação do PDF: {e}")

# Função para copiar nomes de ficheiros para Excel
def copiar_nomes_para_excel(parent=None):
    try:
        # Importa openpyxl aqui para garantir que está instalado
        import openpyxl
        from openpyxl import Workbook, load_workbook
        
        excel_window = Toplevel(parent)
        apply_icon(excel_window)  # Aplica o ícone
        excel_window.title("Copiar Nomes para Excel")
        center_window(excel_window, 700, 500)  # Aumentado para o dobro (era 700x550)
        excel_window.configure(bg="#F0F0F0")
        
        main_frame = Frame(excel_window, bg="#F0F0F0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Título
        title_label = Label(main_frame, text="Copiar Nomes de Ficheiros para Excel", 
                           font=("Arial", 16, "bold"), bg="#F0F0F0")
        title_label.grid(row=0, column=0, columnspan=3, pady=15, sticky="w")
        
        # Entrada para pasta
        Label(main_frame, text="Pasta com ficheiros:", font=("Arial", 11), bg="#F0F0F0").grid(row=1, column=0, sticky="w", pady=10)
        pasta_var = StringVar()
        entrada_pasta = Entry(main_frame, textvariable=pasta_var, width=50, font=("Arial", 10))
        entrada_pasta.grid(row=1, column=1, padx=10, pady=10)
        
        def selecionar_pasta():
            pasta = filedialog.askdirectory(title="Selecione a pasta com os ficheiros")
            if pasta:
                pasta_var.set(pasta)
                
        Button(main_frame, text="Procurar", command=selecionar_pasta, 
              font=("Arial", 10), bg="#4CAF50", fg="white", width=10).grid(row=1, column=2, padx=10, pady=10)
        
        # Entrada para ficheiro Excel
        Label(main_frame, text="Ficheiro Excel:", font=("Arial", 11), bg="#F0F0F0").grid(row=2, column=0, sticky="w", pady=10)
        excel_var = StringVar()
        entrada_excel = Entry(main_frame, textvariable=excel_var, width=50, font=("Arial", 10))
        entrada_excel.grid(row=2, column=1, padx=10, pady=10)
        
        def selecionar_excel():
            excel = filedialog.asksaveasfilename(
                title="Guardar em ficheiro Excel",
                defaultextension=".xlsx",
                filetypes=[("Ficheiros Excel", "*.xlsx")]
            )
            if excel:
                excel_var.set(excel)
        
        Button(main_frame, text="Procurar", command=selecionar_excel, 
              font=("Arial", 10), bg="#4CAF50", fg="white", width=10).grid(row=2, column=2, padx=10, pady=10)
        
        # Entrada para coluna e linha
        param_frame = Frame(main_frame, bg="#F0F0F0")
        param_frame.grid(row=3, column=0, columnspan=3, sticky="w", pady=10)
        
        Label(param_frame, text="Coluna inicial:", font=("Arial", 11), bg="#F0F0F0").grid(row=0, column=0, sticky="w", padx=10)
        coluna_var = StringVar(value="A")
        entrada_coluna = Entry(param_frame, textvariable=coluna_var, width=5, font=("Arial", 10))
        entrada_coluna.grid(row=0, column=1, padx=10)
        
        Label(param_frame, text="Linha inicial:", font=("Arial", 11), bg="#F0F0F0").grid(row=0, column=2, sticky="w", padx=10)
        linha_var = StringVar(value="1")
        entrada_linha = Entry(param_frame, textvariable=linha_var, width=5, font=("Arial", 10))
        entrada_linha.grid(row=0, column=3, padx=10)
        
        # Botão executar
        def executar_copia():
            pasta = pasta_var.get()
            ficheiro_excel = excel_var.get()
            coluna = coluna_var.get().strip().upper()
            linha_inicial = linha_var.get().strip()
            
            if not pasta or not ficheiro_excel or not coluna or not linha_inicial:
                messagebox.showerror("Erro", "Por favor, preencha todos os campos.")
                return
                
            if not coluna.isalpha() or len(coluna) > 3:
                messagebox.showerror("Erro", "Coluna inválida. Insira uma letra ou combinação de letras válida.")
                return
                
            try:
                linha_inicial = int(linha_inicial)
                if linha_inicial <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Erro", "Linha inicial deve ser um número positivo.")
                return
            
            progress_window, progress_bar, progress_var = show_progress_window("A processar ficheiros", excel_window)
            
            def processar_ficheiros():
                try:
                    nomes_ficheiros = sorted(os.listdir(pasta))
                    total_files = len(nomes_ficheiros)
                    
                    try:
                        wb = load_workbook(ficheiro_excel)
                        ws = wb.active
                    except FileNotFoundError:
                        wb = Workbook()
                        ws = wb.active
                    
                    for i, nome in enumerate(nomes_ficheiros):
                        # Escreve o nome sem extensão na célula
                        ws[f"{coluna}{linha_inicial + i}"] = os.path.splitext(nome)[0]
                        
                        # Atualiza progresso
                        valor = int((i + 1) / total_files * 100)
                        progress_bar["value"] = valor
                        progress_var.set(f"A processar: {valor}%")
                        progress_window.update()
                    
                    wb.save(ficheiro_excel)
                    
                    progress_bar["value"] = 100
                    progress_var.set("Concluído: 100%")
                    progress_window.update()
                    time.sleep(1)
                    progress_window.destroy()
                    
                    messagebox.showinfo("Concluído", f"Nomes guardados em:\n{ficheiro_excel}")
                    abrir_pasta(ficheiro_excel)
                    
                except Exception as e:
                    progress_window.destroy()
                    messagebox.showerror("Erro", f"Erro ao processar ficheiros: {e}")
            
            threading.Thread(target=processar_ficheiros, daemon=True).start()
            
        button_frame = Frame(main_frame, bg="#F0F0F0")
        button_frame.grid(row=4, column=0, columnspan=3, pady=20)
        
        Button(button_frame, text="Executar", command=executar_copia, 
              font=("Arial", 11, "bold"), bg="#007BFF", fg="white", width=15, height=1).pack(side="left", padx=10)
        
        Button(button_frame, text="Voltar", command=excel_window.destroy, 
              font=("Arial", 11), bg="#6C757D", fg="white", width=15, height=1).pack(side="left", padx=10)
        
        # Disclaimers
        disclaimer_frame = Frame(main_frame, bg="#F0F0F0", pady=10)
        disclaimer_frame.grid(row=5, column=0, columnspan=3, sticky="w")
        
        Label(disclaimer_frame, text="Notas importantes:", font=("Arial", 10, "bold"), bg="#F0F0F0").pack(anchor="w")
        
        disclaimers = [
            "- O ficheiro Excel deve estar fechado durante a operação.",
            "- Se o programa perguntar se é para substituir o ficheiro, clique em 'Sim'.",
            "- O programa apenas guardará os nomes sem extensão dos ficheiros."
        ]
        
        for disclaimer in disclaimers:
            Label(disclaimer_frame, text=disclaimer, font=("Arial", 9), bg="#F0F0F0", justify="left").pack(anchor="w", pady=2)
            
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar a cópia de nomes: {e}")

# Função para comprimir ficheiros
def comprimir_ficheiro(parent=None):
    try:
        file_path = filedialog.askopenfilename(
            parent=parent,
            title="Selecione o ficheiro para comprimir"
        )
        if not file_path:
            return
            
        file_name, file_ext = os.path.splitext(os.path.basename(file_path))
        dir_path = os.path.dirname(file_path)
        
        # Sugerir pasta de destino e nome de ficheiro
        save_path = filedialog.asksaveasfilename(
            parent=parent,
            title="Guardar ficheiro comprimido",
            initialdir=dir_path,
            initialfile=f"{file_name}_compressed{file_ext}",
            defaultextension=file_ext
        )
        if not save_path:
            return
        
        progress_window, progress_bar, progress_var = show_progress_window("A comprimir ficheiro", parent)
        
        def executar_compressao():
            try:
                # Definir resultado_save_path para guardar o caminho final
                resultado_save_path = save_path
                
                # Atualiza progresso
                progress_bar["value"] = 10
                progress_var.set("A iniciar compressão: 10%")
                progress_window.update()
                
                # Verifica extensão e aplica método de compressão adequado
                if file_ext.lower() == ".pdf":
                    import fitz
                    
                    # Atualiza progresso
                    progress_bar["value"] = 30
                    progress_var.set("A comprimir PDF: 30%")
                    progress_window.update()
                    
                    doc = fitz.open(file_path)
                    
                    # Atualiza progresso
                    progress_bar["value"] = 50
                    progress_var.set("A processar páginas: 50%")
                    progress_window.update()
                    
                    # Salvar com compressão
                    doc.save(
                        resultado_save_path,
                        garbage=4,
                        deflate=True,
                        clean=True
                    )
                    doc.close()
                    
                elif file_ext.lower() in [".jpg", ".jpeg", ".png"]:
                    from PIL import Image
                    
                    # Atualiza progresso
                    progress_bar["value"] = 30
                    progress_var.set("A comprimir imagem: 30%")
                    progress_window.update()
                    
                    img = Image.open(file_path)
                    
                    # Atualiza progresso
                    progress_bar["value"] = 50
                    progress_var.set("A processar imagem: 50%")
                    progress_window.update()
                    
                    # Salvar com qualidade reduzida mas preservando detalhes
                    img.save(resultado_save_path, optimize=True, quality=85)
                    
                else:
                    # Para outros tipos de ficheiro, usar compressão ZIP
                    import zipfile
                    
                    # Atualiza progresso
                    progress_bar["value"] = 30
                    progress_var.set("A comprimir ficheiro: 30%")
                    progress_window.update()
                    
                    # Criar nome para o arquivo ZIP
                    zip_path = os.path.splitext(resultado_save_path)[0] + ".zip"
                    
                    # Comprimir o ficheiro
                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(file_path, os.path.basename(file_path))
                        
                    # Atualiza progresso
                    progress_bar["value"] = 70
                    progress_var.set("A finalizar compressão: 70%")
                    progress_window.update()
                    
                    # Atualiza o caminho do arquivo final para o ZIP
                    resultado_save_path = zip_path
                
                # Finaliza progresso
                progress_bar["value"] = 100
                progress_var.set("Concluído: 100%")
                progress_window.update()
                time.sleep(1)
                progress_window.destroy()
                
                # Calcula taxas de compressão
                original_size = os.path.getsize(file_path)
                compressed_size = os.path.getsize(resultado_save_path)
                reduction = (1 - (compressed_size / original_size)) * 100
                
                messagebox.showinfo(
                    "Compressão concluída", 
                    f"Ficheiro comprimido guardado em:\n{resultado_save_path}\n\n"
                    f"Tamanho original: {original_size/1024:.1f} KB\n"
                    f"Tamanho comprimido: {compressed_size/1024:.1f} KB\n"
                    f"Redução: {reduction:.1f}%"
                )
                abrir_pasta(resultado_save_path)
                
            except Exception as e:
                progress_window.destroy()
                messagebox.showerror("Erro", f"Erro ao comprimir ficheiro: {e}")
        
        threading.Thread(target=executar_compressao, daemon=True).start()
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar a compressão: {e}")

# Menu de trabalho com PDF
def menu_pdf(parent=None):
    pdf_window = Toplevel(parent)
    apply_icon(pdf_window)  # Aplica o ícone
    pdf_window.title("Trabalhar com PDF")
    center_window(pdf_window, 700, 500)  # Aumentado para o dobro (era 600x400)
    pdf_window.configure(bg="#F0F0F0")
    
    main_frame = Frame(pdf_window, bg="#F0F0F0", padx=30, pady=30)
    main_frame.pack(fill="both", expand=True)
    
    # Título
    title_label = Label(main_frame, text="Selecione uma operação com PDF", 
                       font=("Arial", 16, "bold"), bg="#F0F0F0")
    title_label.pack(pady=20)
    
    # Botões
    button_style = {"font": ("Arial", 12), "width": 20, "height": 2}
    
    Button(main_frame, text="Juntar PDFs", command=lambda: juntar_pdfs(pdf_window), 
          bg="#3498DB", fg="white", **button_style).pack(pady=10)
    
    Button(main_frame, text="Separar PDF", command=lambda: separar_pdf(pdf_window), 
          bg="#3498DB", fg="white", **button_style).pack(pady=10)
    
    Button(main_frame, text="Voltar", command=pdf_window.destroy, 
          bg="#6C757D", fg="white", **button_style).pack(pady=20)

# Menu de trabalho com Excel
def menu_excel(parent=None):
    excel_window = Toplevel(parent)
    apply_icon(excel_window)  # Aplica o ícone
    excel_window.title("Trabalhar com Excel")
    center_window(excel_window, 700, 500)  # Aumentado para o dobro (era 600x400)
    excel_window.configure(bg="#F0F0F0")
    
    main_frame = Frame(excel_window, bg="#F0F0F0", padx=30, pady=30)
    main_frame.pack(fill="both", expand=True)
    
    # Título
    title_label = Label(main_frame, text="Selecione uma operação com Excel", 
                       font=("Arial", 16, "bold"), bg="#F0F0F0")
    title_label.pack(pady=20)
    
    # Botões
    button_style = {"font": ("Arial", 12), "width": 30, "height": 2}
    
    Button(main_frame, text="Copiar nomes de ficheiros para Excel", 
          command=lambda: copiar_nomes_para_excel(excel_window), 
          bg="#27AE60", fg="white", **button_style).pack(pady=10)
    
    Button(main_frame, text="Voltar", command=excel_window.destroy, 
          bg="#6C757D", fg="white", **button_style).pack(pady=20)

# Menu Principal
def menu_principal():
    try:
        # Verifica se já existe uma instância rodando
        check_single_instance()
        
        # Inicializa a janela principal
        root = Tk()
        
        # Define um tamanho mínimo para a janela principal
        root.minsize(700, 500)
        
        # Aplica o ícone à janela principal
        apply_icon(root)
        
        # Verifica atualizações e instala dependências
        check_and_install_dependencies(root)
        
        # Esconde a janela principal temporariamente durante o splash
        root.withdraw()
        
        # Mostra a tela de splash
        show_splash_screen(root)
        
        # Configura a janela principal antes de mostrá-la
        root.title("TirelessBud - Ferramentas de Produtividade")
        
        # Define o tamanho explicitamente antes de exibir
        root.geometry("900x600")  # Tamanho aumentado para garantir visibilidade
        
        # Configura a aparência
        root.configure(bg="#F0F0F0")
        
        # Centraliza a janela na tela
        center_window(root, 900, 600)
        
        # Mostra a janela principal
        root.deiconify()
        
        main_frame = Frame(root, bg="#F0F0F0", padx=40, pady=40)
        main_frame.pack(fill="both", expand=True)
        
        # Título
        title_label = Label(main_frame, text="Tipo de ficheiro a trabalhar?", 
                           font=("Arial", 18, "bold"), bg="#F0F0F0")
        title_label.pack(pady=20)
        
        # Botões
        button_style = {"font": ("Arial", 14), "width": 25, "height": 2}
        
        Button(main_frame, text="PDF", command=lambda: menu_pdf(root), 
              bg="#E74C3C", fg="white", **button_style).pack(pady=10)
        
        Button(main_frame, text="Excel", command=lambda: menu_excel(root), 
              bg="#27AE60", fg="white", **button_style).pack(pady=10)
        
        Button(main_frame, text="Diminuir tamanho do ficheiro", command=lambda: comprimir_ficheiro(root), 
              bg="#3498DB", fg="white", **button_style).pack(pady=10)
        
        # Rodapé
        footer = Label(root, text="TirelessBud v" + VERSION + " | github.com/joaoafs", 
                      font=("Arial", 9, "italic"), fg="#666666", bg="#F0F0F0")
        footer.pack(side="bottom", pady=10)
        
        # Executa a janela
        root.mainloop()
        
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro inesperado: {str(e)}")
        sys.exit(1)

# Inicialização
if __name__ == "__main__":
    menu_principal()